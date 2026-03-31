import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

hdr_font = Font(name="JetBrains Mono", bold=True, size=9, color="FFFFFF")
body_font = Font(name="Arial", size=9, color="000000")
wrap = Alignment(wrap_text=True, vertical="top")
center_wrap = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"), top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))

provider_colors = {
    "Anthropic": "991B1B",
    "OpenAI": "1E40AF",
    "Google": "065F46",
    "DeepSeek": "5B21B6",
}

ws = wb.active
ws.title = "ALL MODELS"
ws.sheet_properties.tabColor = "3B82F6"

# Column widths
col_widths = [22, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[chr(64+i) if i <= 26 else "A"].width = w

ws.column_dimensions["A"].width = 22
for col_letter in ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P"]:
    ws.column_dimensions[col_letter].width = 15

# Title
ws.merge_cells("A1:P1")
ws.cell(row=1, column=1, value="ALL MODEL SETTINGS COMPARISON -- MARCH 2026").font = Font(name="JetBrains Mono", size=12, bold=True, color="3B82F6")
ws.cell(row=1, column=1).fill = PatternFill(start_color="050810", end_color="050810", fill_type="solid")
ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Provider header row
row = 2
providers_row = [
    ("", ""),
    ("ANTHROPIC", "991B1B"), ("ANTHROPIC", "991B1B"), ("ANTHROPIC", "991B1B"), ("ANTHROPIC", "991B1B"),
    ("OPENAI", "1E40AF"), ("OPENAI", "1E40AF"), ("OPENAI", "1E40AF"), ("OPENAI", "1E40AF"),
    ("GOOGLE", "065F46"), ("GOOGLE", "065F46"), ("GOOGLE", "065F46"), ("GOOGLE", "065F46"),
    ("DEEPSEEK", "5B21B6"), ("DEEPSEEK", "5B21B6"), ("DEEPSEEK", "5B21B6"),
]
for i, (name, color) in enumerate(providers_row, 1):
    cell = ws.cell(row=row, column=i, value=name)
    cell.font = Font(name="JetBrains Mono", size=8, bold=True, color="FFFFFF")
    if color:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    else:
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    cell.alignment = center_wrap
    cell.border = thin

# Merge provider headers
ws.merge_cells("B2:E2")
ws.merge_cells("F2:I2")
ws.merge_cells("J2:M2")
ws.merge_cells("N2:P2")

# Model header row
row = 3
models = [
    "SETTING",
    "Opus 4.6", "Sonnet 4.6", "Haiku 4.5", "Sonnet 4.5",
    "GPT-5.4", "GPT-5.4 Pro", "GPT-5 mini", "o3",
    "Gemini 2.5 Pro", "Gemini 2.5 Flash", "Gemini 3.1 Pro", "Gemini 2.0 Flash",
    "DeepSeek V3.2", "DeepSeek R1", "DeepSeek V3.1",
]
model_colors = ["1E293B"] + ["991B1B"]*4 + ["1E40AF"]*4 + ["065F46"]*4 + ["5B21B6"]*3

for i, (m, c) in enumerate(zip(models, model_colors), 1):
    cell = ws.cell(row=row, column=i, value=m)
    cell.font = hdr_font
    cell.fill = PatternFill(start_color=c, end_color=c, fill_type="solid")
    cell.alignment = center_wrap
    cell.border = thin
ws.row_dimensions[row].height = 35
ws.freeze_panes = "B4"

# Data rows
data = [
    # PRICING
    ("PRICING & SPECS", None),
    ("Input $/MTok", "$5.00", "$3.00", "$1.00", "$3.00", "$2.50", "$10.00", "$0.40", "$2.00", "$1.25", "$0.30", "Preview", "$0.10", "$0.28", "$0.56", "$0.28"),
    ("Output $/MTok", "$25.00", "$15.00", "$5.00", "$15.00", "$15.00", "$30.00", "$1.60", "$8.00", "$10.00", "$2.50", "Preview", "$0.40", "$1.12", "$1.68", "$0.42"),
    ("Context Window", "1M (beta)", "1M (beta)", "200K", "1M", "1.05M", "1.05M", "1.05M", "200K", "1M", "1M", "1M", "1M", "128K", "128K", "128K"),
    ("Max Output Tokens", "128K", "64K", "8,192", "64K", "128K", "128K", "128K", "100K", "65K", "65K", "65K", "8,192", "8K", "64K", "8K"),
    ("Release Date", "Feb 5 2026", "Feb 17 2026", "Oct 2025", "Sep 2025", "Mar 5 2026", "Mar 5 2026", "Aug 2025", "Apr 2025", "Jun 2025", "Jun 2025", "Mar 2026", "Feb 2025", "Mar 2026", "Jan 2025", "Jun 2026"),
    ("Open Source", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "Yes (MIT)", "Yes (MIT)", "Yes (MIT)"),

    # MODEL PARAMETERS
    ("MODEL PARAMETERS", None),
    ("temperature range", "0.0-1.0", "0.0-1.0", "0.0-1.0", "0.0-1.0", "0.0-2.0", "0.0-2.0", "0.0-2.0", "0.0-2.0", "0.0-2.0", "0.0-2.0", "0.0-2.0", "0.0-2.0", "0.0-2.0", "0.0-2.0", "0.0-2.0"),
    ("temperature default", "1.0", "1.0", "1.0", "1.0", "1.0", "1.0", "1.0", "1.0", "1.0", "1.0", "1.0", "1.0", "1.0", "1.0", "1.0"),
    ("top_p", "Yes (exclusive w/ temp)", "Yes (exclusive w/ temp)", "Yes", "Yes (exclusive w/ temp)", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes"),
    ("top_k", "Not exposed", "Not exposed", "Not exposed", "Not exposed", "Not exposed", "Not exposed", "Not exposed", "Not exposed", "Yes (1-100)", "Yes (1-100)", "Yes (1-100)", "Yes (1-100)", "Not exposed", "Not exposed", "Not exposed"),
    ("temp + top_p together", "REJECTED", "REJECTED", "Allowed", "REJECTED", "Allowed", "Allowed", "Allowed", "Allowed", "Allowed", "Allowed", "Allowed", "Allowed", "Allowed", "Allowed", "Allowed"),
    ("frequency_penalty", "No", "No", "No", "No", "Yes (-2 to 2)", "Yes (-2 to 2)", "Yes (-2 to 2)", "Yes (-2 to 2)", "Yes", "Yes", "Yes", "Yes", "Yes (-2 to 2)", "Yes (-2 to 2)", "Yes (-2 to 2)"),
    ("presence_penalty", "No", "No", "No", "No", "Yes (-2 to 2)", "Yes (-2 to 2)", "Yes (-2 to 2)", "Yes (-2 to 2)", "No", "No", "No", "No", "Yes (-2 to 2)", "Yes (-2 to 2)", "Yes (-2 to 2)"),
    ("seed (determinism)", "No", "No", "No", "No", "Yes (best effort)", "Yes (best effort)", "Yes (best effort)", "Yes (best effort)", "No", "No", "No", "No", "Yes", "Yes", "Yes"),

    # REASONING / THINKING
    ("REASONING / THINKING", None),
    ("Extended Thinking", "Adaptive (default)", "Adaptive (default)", "Yes (enabled)", "Yes (enabled)", "Yes (Thinking)", "Yes (Thinking)", "Yes (Thinking)", "Yes (native)", "Yes (thinking)", "Yes (thinking)", "Yes (thinking)", "No", "No", "Yes (CoT native)", "Yes (Deep Thinking)"),
    ("Effort Control", "low/med/high/max", "low/med/high/max", "No", "No", "none/low/med/high/xhigh", "none/low/med/high/xhigh", "none/low/med/high/xhigh", "low/med/high", "thinkingBudget", "thinkingBudget", "thinkingBudget", "No", "No", "No", "mode param"),
    ("budget_tokens", "Deprecated", "Deprecated", "Yes (min 1024)", "Yes (min 1024)", "N/A (effort)", "N/A (effort)", "N/A (effort)", "N/A", "Yes (0-24576)", "Yes (0-24576)", "TBD", "No", "No", "No", "No"),
    ("Thinking token billing", "Output rate", "Output rate", "Output rate", "Output rate", "Reasoning rate", "Reasoning rate", "Reasoning rate", "Reasoning rate", "Thinking rate", "Thinking rate", "TBD", "N/A", "N/A", "Output rate", "Output rate"),
    ("Interleaved thinking", "Auto (adaptive)", "Beta header", "No", "No", "N/A", "N/A", "N/A", "N/A", "No", "No", "No", "No", "No", "No", "No"),

    # TOOLS & CAPABILITIES
    ("TOOLS & CAPABILITIES", None),
    ("Web Search", "Yes ($10/1K)", "Yes ($10/1K)", "Yes ($10/1K)", "Yes ($10/1K)", "Yes ($30/1K)", "Yes ($30/1K)", "Yes ($30/1K)", "Yes", "Yes (Grounding)", "Yes (Grounding)", "Yes (Grounding)", "Yes (Grounding)", "No", "No", "No"),
    ("Code Execution", "Yes", "Yes", "Yes", "Yes", "Yes (container)", "Yes (container)", "Yes (container)", "No", "Yes", "Yes", "Yes", "Yes", "No", "No", "No"),
    ("Computer Use", "Yes (beta)", "Yes (beta)", "No", "No", "Yes", "Yes", "No", "No", "No", "No", "No", "No", "No", "No", "No"),
    ("Tool / Function Call", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes"),
    ("Image Input", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "No", "No", "No"),
    ("Image Generation", "No", "No", "No", "No", "Yes (GPT Image)", "Yes (GPT Image)", "No", "No", "Yes (Imagen)", "Yes (Imagen)", "Yes (Imagen)", "No", "No", "No", "No"),
    ("Video Input", "No", "No", "No", "No", "No", "No", "No", "No", "Yes", "Yes", "Yes", "Yes", "No", "No", "No"),
    ("Audio Input", "No", "No", "No", "No", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "No", "No", "No"),
    ("Structured Output", "Yes (GA)", "Yes (GA)", "Yes (GA)", "Yes (GA)", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "No", "No", "No"),
    ("Batch API", "Yes (50% off)", "Yes (50% off)", "Yes (50% off)", "Yes (50% off)", "Yes (50% off)", "Yes (50% off)", "Yes (50% off)", "No", "Yes", "Yes", "TBD", "Yes", "Yes", "Yes", "Yes"),
    ("Prompt Caching", "Yes (90% save)", "Yes (90% save)", "Yes (90% save)", "Yes (90% save)", "Yes (50% save)", "Yes (50% save)", "Yes (50% save)", "No", "Yes", "Yes", "TBD", "Yes", "Yes (75% save)", "Yes (75% save)", "Yes (75% save)"),

    # SPEED & INFRA
    ("SPEED & INFRASTRUCTURE", None),
    ("Fast Mode", "Yes ($30/$150)", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No"),
    ("Data Residency", "Yes (US, 1.1x)", "Yes (US, 1.1x)", "No", "No", "Yes (10% uplift)", "Yes (10% uplift)", "Yes (10% uplift)", "No", "Yes (regional)", "Yes (regional)", "Yes (regional)", "Yes (regional)", "No", "No", "No"),
    ("Context Compaction", "Yes (server)", "Yes (server)", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No"),
    ("Streaming", "Yes (required >64K)", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes"),
    ("Self-hosted option", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "No", "Yes", "Yes", "Yes"),

    # BENCHMARKS (approximate as of March 2026)
    ("KEY BENCHMARKS", None),
    ("SWE-bench Verified", "80.8%", "79.6%", "73.3%", "77.2%", "~80%", "~82%", "N/A", "72%", "~65%", "~55%", "TBD", "~45%", "~50%", "~55%", "~52%"),
    ("HLE (Humanity Last Exam)", "53%", "52%", "N/A", "N/A", "~50%", "~55%", "N/A", "~42%", "~45%", "~35%", "TBD", "N/A", "N/A", "~40%", "N/A"),
    ("GPQA Diamond", "~75%", "~70%", "~55%", "~65%", "~75%", "~80%", "N/A", "~80%", "~70%", "~60%", "TBD", "~50%", "~60%", "~72%", "~62%"),
    ("OSWorld (Computer Use)", "N/A", "72.5%", "N/A", "50.7%", "75%", "75%", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"),
]

row = 4
for d in data:
    if d[1] is None:
        # Category header
        for col in range(1, 17):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            cell.border = thin
        ws.cell(row=row, column=1, value=d[0]).font = Font(name="JetBrains Mono", bold=True, size=10, color="3B82F6")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        ws.row_dimensions[row].height = 22
        row += 1
        continue

    vals = list(d)
    for col_idx in range(1, min(17, len(vals)+1)):
        cell = ws.cell(row=row, column=col_idx, value=vals[col_idx-1])
        if col_idx == 1:
            cell.font = Font(name="JetBrains Mono", size=9, bold=True, color="000000")
        else:
            cell.font = body_font
        cell.alignment = center_wrap if col_idx > 1 else wrap
        cell.border = thin

        # Color code important values
        val = str(vals[col_idx-1]).upper() if col_idx > 1 and col_idx-1 < len(vals) else ""
        if val == "REJECTED":
            cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            cell.font = Font(name="Arial", size=9, color="991B1B", bold=True)
        elif val.startswith("YES"):
            cell.font = Font(name="Arial", size=9, color="065F46")
        elif val == "NO" or val == "N/A":
            cell.font = Font(name="Arial", size=9, color="991B1B")
        elif val == "DEPRECATED":
            cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            cell.font = Font(name="Arial", size=9, color="92400E", bold=True)

    ws.row_dimensions[row].height = 35
    row += 1

# Notes row
row += 1
ws.merge_cells(f"A{row}:P{row}")
ws.cell(row=row, column=1, value="NOTES: Prices as of March 2026. Benchmarks are approximate and may vary by evaluation method. 'Preview' = pricing not yet finalized. DeepSeek prices shown are direct API (platform.deepseek.com); prices vary by provider (Groq, Fireworks, etc). Gemini 3.1 Pro is in preview. GPT-5.4 context >272K input tokens charged at 2x input / 1.5x output. Claude 4.6 models reject temperature + top_p together (breaking change since Opus 4.1). DeepSeek V4 expected mid-2026.").font = Font(name="Arial", size=8, color="666666", italic=True)
ws.cell(row=row, column=1).alignment = wrap
ws.row_dimensions[row].height = 50

output_path = "/mnt/user-data/outputs/ALL_MODEL_SETTINGS_COMPARISON.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Models: 15")
print(f"Settings rows: {row - 4}")
