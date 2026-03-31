import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ALL MODELS"
ws.sheet_properties.tabColor = "3B82F6"

hdr_font = Font(name="JetBrains Mono", bold=True, size=9, color="FFFFFF")
hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
prov_font = Font(name="JetBrains Mono", bold=True, size=10, color="FFFFFF")
body_font = Font(name="Arial", size=9, color="000000")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin = Border(left=Side(style="thin",color="D1D5DB"), right=Side(style="thin",color="D1D5DB"),
              top=Side(style="thin",color="D1D5DB"), bottom=Side(style="thin",color="D1D5DB"))

provider_fills = {
    "ANTHROPIC": PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid"),
    "OPENAI": PatternFill(start_color="059669", end_color="059669", fill_type="solid"),
    "GOOGLE": PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid"),
    "DEEPSEEK": PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid"),
}

# Headers
headers = [
    "PROVIDER", "MODEL", "MODEL STRING", "CONTEXT WINDOW", "MAX OUTPUT",
    "TEMPERATURE", "TOP_P", "TOP_K", "TEMP+TOP_P TOGETHER?",
    "EXTENDED THINKING", "EFFORT LEVELS", "REASONING EFFORT",
    "FREQUENCY PENALTY", "PRESENCE PENALTY", "VERBOSITY CONTROL",
    "STRUCTURED OUTPUT", "VISION (IMAGE INPUT)", "COMPUTER USE",
    "TOOL USE", "WEB SEARCH", "AGENT TEAMS",
    "INPUT $/MTok", "OUTPUT $/MTok", "FAST MODE",
    "COMPACTION", "1M CONTEXT BETA", "NOTES"
]

for i, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = thin

ws.freeze_panes = "D2"

# Column widths
widths = [12,22,28,12,12,12,12,8,14,14,14,14,10,10,12,10,10,10,10,10,10,10,10,10,10,10,35]
for i, w in enumerate(widths):
    ws.column_dimensions[chr(65+i) if i < 26 else "A" + chr(65+i-26)].width = w

# ─── MODEL DATA ───
models = [
    # ANTHROPIC
    ("ANTHROPIC", "Opus 4.6", "claude-opus-4-6", "200K (1M beta)", "128K", "0.0-1.0 (default 1.0)", "0.0-1.0", "N/A", "NO (since 4.1)", "YES (adaptive)", "low/med/high/max", "N/A", "NO", "NO", "NO", "YES (GA)", "YES", "YES", "YES", "YES ($10/1K)", "YES", "$5", "$25", "YES ($30/$150)", "YES (server-side)", "YES (no surcharge)", "Flagship. Best reasoning. Agent Teams. budget_tokens deprecated, use adaptive thinking."),
    ("ANTHROPIC", "Opus 4.6 Fast", "claude-opus-4-6 +speed=fast", "200K (1M beta)", "128K", "0.0-1.0 (default 1.0)", "0.0-1.0", "N/A", "NO", "YES (adaptive)", "low/med/high/max", "N/A", "NO", "NO", "NO", "YES", "YES", "YES", "YES", "YES", "YES", "$30", "$150", "IS fast mode", "YES", "YES", "2.5x faster output. Same intelligence. 6x cost."),
    ("ANTHROPIC", "Sonnet 4.6", "claude-sonnet-4-6", "200K (1M beta)", "64K", "0.0-1.0 (default 1.0)", "0.0-1.0", "N/A", "NO (since 4.1)", "YES (adaptive)", "low/med/high/max", "N/A", "NO", "NO", "NO", "YES (GA)", "YES", "YES", "YES", "YES", "NO", "$3", "$15", "NO", "YES", "YES (no surcharge)", "Workhorse. 80% of Opus quality at 20% cost. Effort default: medium recommended."),
    ("ANTHROPIC", "Sonnet 5", "claude-sonnet-5-*", "1M (default)", "64K", "0.0-1.0", "0.0-1.0", "N/A", "NO", "YES", "TBD", "N/A", "NO", "NO", "NO", "YES", "YES", "YES", "YES", "YES", "Dev Team", "TBD", "TBD", "NO", "YES", "YES (default)", "Latest gen. Dev Team (multi-agent). Pricing TBD at time of writing."),
    ("ANTHROPIC", "Haiku 4.5", "claude-haiku-4-5-20251001", "200K", "8K", "0.0-1.0 (default 1.0)", "0.0-1.0", "N/A", "NO (since 4.1)", "YES (manual)", "N/A", "N/A", "NO", "NO", "NO", "YES (GA)", "YES", "NO", "YES", "NO", "NO", "$1", "$5", "NO", "NO", "NO", "Fastest/cheapest. First Haiku with extended thinking. 73.3% SWE-bench."),

    # OPENAI
    ("OPENAI", "GPT-5.4", "gpt-5.4", "1M (~1,050K)", "128K (Pro)", "YES (0.0-2.0)", "YES (0.0-1.0)", "N/A", "YES (both OK)", "Built-in reasoning", "N/A", "low/med/high", "YES (-2 to 2)", "YES (-2 to 2)", "low/med/high", "YES (json_schema)", "YES", "YES (built-in)", "YES + tool_search", "NO (native)", "NO", "TBD", "TBD", "NO", "YES (native)", "YES (default)", "Latest flagship. Verbosity param replaces conciseness hacks. Computer use built-in. Compaction native."),
    ("OPENAI", "GPT-5.4 Pro", "gpt-5.4-pro-2026-03-05", "1M", "128K", "YES (0.0-2.0)", "YES", "N/A", "YES", "Deep reasoning", "N/A", "low/med/high/xhigh", "YES", "YES", "low/med/high", "YES", "YES", "YES", "YES", "NO", "NO", "TBD", "TBD", "NO", "YES", "YES", "Higher compute variant. xhigh effort = most expensive but most capable."),
    ("OPENAI", "GPT-5.2", "gpt-5.2", "400K", "128K", "NOT SUPPORTED", "YES (0.0-1.0)", "N/A", "N/A (temp blocked)", "Reasoning model", "N/A", "low/med/high", "NO", "NO", "NO", "YES", "YES", "NO", "YES", "NO", "NO", "$2.50", "$10", "NO", "NO", "NO", "Reasoning model. Temperature REJECTED by API. Use reasoning_effort instead."),
    ("OPENAI", "GPT-4o", "gpt-4o", "128K", "16K", "YES (0.0-2.0)", "YES (0.0-1.0)", "N/A", "YES (both OK)", "NO", "N/A", "N/A", "YES (-2 to 2)", "YES (-2 to 2)", "NO", "YES", "YES", "NO", "YES", "NO", "NO", "$2.50", "$10", "NO", "NO", "NO", "Non-reasoning. Full parameter control. Freq/presence penalties available. Being phased out."),
    ("OPENAI", "GPT-4o-mini", "gpt-4o-mini", "128K", "16K", "YES (0.0-2.0)", "YES", "N/A", "YES", "NO", "N/A", "N/A", "YES", "YES", "NO", "YES", "YES", "NO", "YES", "NO", "NO", "$0.15", "$0.60", "NO", "NO", "NO", "Cheapest OpenAI. Good for routing/classification."),

    # GOOGLE
    ("GOOGLE", "Gemini 3.1 Pro Preview", "gemini-3.1-pro-preview", "1M+", "65K", "YES (0.0-2.0)", "YES (0.0-1.0)", "Fixed at 64", "YES (both OK)", "YES (thinking)", "N/A", "N/A", "NO", "NO", "NO", "YES", "YES", "NO", "YES", "Grounding w/ Search", "NO", "TBD", "TBD", "NO", "NO", "YES (default)", "Top of Intelligence Index (57). Optimized for temp 1.0. Lower temps cause looping."),
    ("GOOGLE", "Gemini 2.5 Pro", "gemini-2.5-pro", "1M", "65K", "YES (0.0-2.0)", "YES", "Fixed at 64", "YES", "YES (thinking)", "N/A", "N/A", "NO", "NO", "NO", "YES", "YES (multimodal)", "NO", "YES", "Grounding", "NO", "$1.25-2.50", "$10-15", "NO", "NO", "YES (default)", "State-of-art reasoning. Multimodal: text+image+video+audio input. top_k fixed at 64."),
    ("GOOGLE", "Gemini 2.5 Flash", "gemini-2.5-flash", "1M", "65K", "YES (0.0-2.0)", "YES", "Fixed at 64", "YES", "YES (configurable)", "N/A", "N/A", "NO", "NO", "NO", "YES", "YES (multimodal)", "NO", "YES", "Grounding", "NO", "$0.15-0.30", "$0.60-2.50", "NO", "NO", "YES (default)", "Best price-performance. Thinking ON/OFF. 220 tok/sec. Multimodal input."),
    ("GOOGLE", "Gemini 2.5 Flash-Lite", "gemini-2.5-flash-lite", "1M", "65K", "YES (0.0-2.0)", "YES", "Fixed at 64", "YES", "YES (budget control)", "N/A", "N/A", "NO", "NO", "NO", "YES", "YES", "NO", "YES", "Grounding", "NO", "$0.075", "$0.30", "NO", "NO", "YES (default)", "Cheapest Gemini. Thinking off by default for speed. 393 tok/sec."),

    # DEEPSEEK
    ("DEEPSEEK", "DeepSeek V4", "deepseek-chat (V4)", "128K", "8K", "YES (0.0-2.0, default 1.0)", "YES (0.0-1.0)", "N/A", "YES (both OK)", "Hybrid (enable_thinking)", "N/A", "N/A", "YES (-2 to 2)", "YES (-2 to 2)", "NO", "YES (json)", "NO", "NO", "YES", "NO", "NO", "$0.30", "$0.50", "NO", "NO", "NO", "Latest flagship. 81% SWE-bench. MoE: 671B total, 37B active. Off-peak discounts available."),
    ("DEEPSEEK", "DeepSeek V3.2", "deepseek-chat (V3.2)", "128K", "8K", "YES (0.0-2.0, default 1.0)", "YES", "N/A", "YES", "Hybrid (enable_thinking)", "N/A", "N/A", "YES", "YES", "NO", "YES", "NO", "NO", "YES", "NO", "NO", "$0.28", "$0.42", "NO", "NO", "NO", "Budget workhorse. Same API as V4 but cheaper. MoE architecture."),
    ("DEEPSEEK", "DeepSeek R1", "deepseek-reasoner", "64K input", "32-64K", "IGNORED (no effect)", "IGNORED", "N/A", "N/A", "Always-on CoT", "N/A", "N/A", "IGNORED", "IGNORED", "NO", "YES", "NO", "NO", "YES", "NO", "NO", "$0.55", "$2.19", "NO", "NO", "NO", "Reasoning specialist. Temp/top_p/penalties accepted but have NO EFFECT. 96% cheaper than o1."),
]

row = 2
current_provider = None
for m in models:
    provider = m[0]
    
    # Provider separator row
    if provider != current_provider:
        if current_provider is not None:
            # blank separator
            row += 1
        current_provider = provider
    
    for col_idx, val in enumerate(m, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = body_font
        cell.alignment = center if col_idx <= 3 or col_idx > 21 else center
        cell.border = thin
        
        # Provider column coloring
        if col_idx == 1:
            cell.font = prov_font
            cell.fill = provider_fills.get(provider, hdr_fill)
        
        # Highlight YES/NO for feature columns
        if col_idx >= 10 and col_idx <= 21:
            if str(val).startswith("YES"):
                cell.font = Font(name="Arial", size=9, color="065F46", bold=True)
            elif str(val).startswith("NO"):
                cell.font = Font(name="Arial", size=9, color="991B1B")
            elif "IGNORED" in str(val) or "NOT SUPPORTED" in str(val):
                cell.font = Font(name="Arial", size=9, color="92400E", bold=True)
    
    ws.row_dimensions[row].height = 50
    row += 1

# ─── KEY DIFFERENCES SHEET ───
ws2 = wb.create_sheet("KEY DIFFERENCES")
ws2.sheet_properties.tabColor = "EF4444"
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 80

diffs = [
    ("CRITICAL DIFFERENCES", None),
    ("Temperature + top_p", "ANTHROPIC (Claude 4.x+): CANNOT send both. API returns 400 error. Use temperature only.\nOPENAI (GPT-4o, GPT-5.4): CAN send both.\nOPENAI (GPT-5.2): Temperature NOT SUPPORTED AT ALL. Reasoning model.\nGOOGLE: CAN send both. top_k fixed at 64.\nDEEPSEEK (R1): Accepts both but IGNORES them. No effect on output."),
    ("Temperature range", "ANTHROPIC: 0.0 to 1.0 only.\nOPENAI: 0.0 to 2.0.\nGOOGLE: 0.0 to 2.0. Optimized for 1.0. Lower causes looping on Gemini 3.\nDEEPSEEK: 0.0 to 2.0 (chat). IGNORED on R1."),
    ("Extended thinking", "ANTHROPIC: Adaptive thinking (type: adaptive). Effort: low/med/high/max. budget_tokens DEPRECATED.\nOPENAI: Built-in reasoning. reasoning_effort: low/med/high. GPT-5.4 Pro adds xhigh.\nGOOGLE: Thinking ON/OFF with configurable budget.\nDEEPSEEK: R1 always thinks. V3.2/V4 hybrid via enable_thinking param."),
    ("Frequency/Presence penalties", "ANTHROPIC: NOT AVAILABLE on any model.\nOPENAI: Available on GPT-4o (-2 to 2). NOT on GPT-5.2.\nGOOGLE: NOT AVAILABLE.\nDEEPSEEK: Available on chat models. IGNORED on R1."),
    ("Verbosity control", "ANTHROPIC: NOT AVAILABLE. Control via prompt instructions only.\nOPENAI: GPT-5.4 has verbosity param (low/med/high). Unique to OpenAI.\nGOOGLE: NOT AVAILABLE.\nDEEPSEEK: NOT AVAILABLE."),
    ("Computer use", "ANTHROPIC: YES. Vision-based screenshot loop.\nOPENAI: GPT-5.4 has BUILT-IN computer use. First mainline model.\nGOOGLE: NO.\nDEEPSEEK: NO."),
    ("Agent teams", "ANTHROPIC: YES (Opus 4.6 in Claude Code). Parallel sub-agents.\nOPENAI: NO (Codex agents are separate product).\nGOOGLE: NO.\nDEEPSEEK: NO."),
    ("Context window (default)", "ANTHROPIC: 200K standard. 1M beta (free, no surcharge).\nOPENAI: GPT-5.4 = 1M. GPT-5.2 = 400K. GPT-4o = 128K.\nGOOGLE: 1M default on all Gemini 2.5+.\nDEEPSEEK: 128K (V3.2/V4). 64K input (R1)."),
    ("Max output tokens", "ANTHROPIC: Opus 128K. Sonnet 64K. Haiku 8K.\nOPENAI: GPT-5.4 128K (Pro). GPT-4o 16K.\nGOOGLE: 65K across all Gemini 2.5.\nDEEPSEEK: 8K optimal (V3/V4). 32-64K (R1 with CoT)."),
    ("Cheapest option", "DEEPSEEK V3.2: $0.28/$0.42 per MTok.\nGOOGLE Flash-Lite: $0.075/$0.30.\nANTHROPIC Haiku: $1/$5.\nOPENAI 4o-mini: $0.15/$0.60."),
    ("Most expensive option", "ANTHROPIC Opus Fast: $30/$150.\nOPENAI GPT-5.4 Pro: TBD (expected highest).\nGOOGLE Gemini 2.5 Pro: $2.50/$15.\nDEEPSEEK R1: $0.55/$2.19 (still 96% cheaper than o1)."),
]

for i, (key, val) in enumerate(diffs, 1):
    if val is None:
        ws2.cell(row=i, column=1, value=key).font = Font(name="JetBrains Mono", bold=True, size=12, color="3B82F6")
        ws2.cell(row=i, column=1).fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        ws2.cell(row=i, column=2).fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        continue
    ws2.cell(row=i, column=1, value=key).font = Font(name="JetBrains Mono", bold=True, size=9)
    ws2.cell(row=i, column=1).alignment = wrap
    ws2.cell(row=i, column=1).border = thin
    ws2.cell(row=i, column=2, value=val).font = body_font
    ws2.cell(row=i, column=2).alignment = wrap
    ws2.cell(row=i, column=2).border = thin
    ws2.row_dimensions[i].height = 80

output = "/mnt/user-data/outputs/CROSS_PROVIDER_MODEL_COMPARISON.xlsx"
wb.save(output)
print(f"Saved: {output}")
print(f"Models: {len(models)}")
print(f"Parameters per model: {len(headers)}")
