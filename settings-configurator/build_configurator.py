import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

# ─── STYLES ───
hdr_font = Font(name="JetBrains Mono", bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
cat_font = Font(name="JetBrains Mono", bold=True, size=11, color="3B82F6")
cat_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
setting_font = Font(name="JetBrains Mono", size=10, bold=True, color="000000")
dropdown_font = Font(name="Arial", size=11, bold=True, color="1E40AF")
dropdown_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
body_font = Font(name="Arial", size=9, color="333333")
score_font = Font(name="JetBrains Mono", size=14, bold=True)
score_label_font = Font(name="JetBrains Mono", size=10, bold=True)
meter_fill_green = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
meter_fill_blue = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
meter_fill_red = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
meter_fill_purple = PatternFill(start_color="5B21B6", end_color="5B21B6", fill_type="solid")
meter_fill_amber = PatternFill(start_color="92400E", end_color="92400E", fill_type="solid")
meter_fill_cyan = PatternFill(start_color="155E75", end_color="155E75", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="center", horizontal="left")
center = Alignment(horizontal="center", vertical="center")
thin = Border(
    left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"),
)

# ═══════════════════════════════════════════
# SHEET 1: CONFIGURATOR
# ═══════════════════════════════════════════
ws = wb.active
ws.title = "CONFIGURATOR"
ws.sheet_properties.tabColor = "3B82F6"

# Column widths
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 14
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 14
ws.column_dimensions["I"].width = 14

# ─── TITLE ───
ws.merge_cells("A1:I1")
ws.cell(row=1, column=1, value="CLAUDE DEFAULT BEHAVIORAL HEURISTICS CONFIGURATOR").font = Font(name="JetBrains Mono", size=14, bold=True, color="3B82F6")
ws.cell(row=1, column=1).fill = PatternFill(start_color="050810", end_color="050810", fill_type="solid")
ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

# ─── HEADERS ROW 2 ───
headers = ["SETTING", "YOUR CHOICE (change me)", "WHAT IT DOES", "SPEED", "QUALITY", "ACCURACY", "COST EFF.", "DEPTH", "CREATIVITY"]
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=i, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center if i > 3 else wrap
    cell.border = thin
ws.row_dimensions[2].height = 25
ws.freeze_panes = "A3"

# ─── SETTINGS DATA ───
# Each setting: (name, options_list, default, description, speed, quality, accuracy, cost, depth, creativity impacts per option)
# Impacts stored as dict: option -> (speed, quality, accuracy, cost, depth, creativity)

settings = [
    ("MODEL PARAMETERS", None, None, None, None),
    ("Temperature", 
     ["0.0", "0.1", "0.2", "0.3", "0.4", "0.5", "0.6", "0.7", "0.8", "0.9", "1.0"],
     "1.0",
     "Randomness of output. 0.0=deterministic, 1.0=creative. Range 0.0-1.0. Cannot combine with top_p on Opus 4.6.",
     {"0.0": (0,10,20,0,0,-10), "0.1": (0,10,18,0,0,-8), "0.2": (0,8,15,0,0,-5), "0.3": (0,8,12,0,0,-3), "0.4": (0,5,8,0,0,0), "0.5": (0,3,5,0,0,3), "0.6": (0,0,2,0,0,5), "0.7": (0,0,0,0,0,8), "0.8": (0,-2,-3,0,0,10), "0.9": (0,-5,-5,0,0,13), "1.0": (0,-5,-8,0,0,15)}),

    ("Effort Level",
     ["low", "medium", "high", "max"],
     "high",
     "Opus 4.6 adaptive thinking depth. Low=skip thinking. Max=deepest reasoning. Thinking tokens billed as output.",
     {"low": (30,-25,-20,30,-30,-10), "medium": (15,-10,-5,15,-10,0), "high": (0,0,0,0,10,5), "max": (-20,15,10,-30,25,10)}),

    ("Model",
     ["Haiku 4.5", "Sonnet 4.6", "Opus 4.6", "Opus 4.6 Fast"],
     "Opus 4.6",
     "Which model runs. Haiku=$1/$5, Sonnet=$3/$15, Opus=$5/$25, Opus Fast=$30/$150 per MTok.",
     {"Haiku 4.5": (40,-30,-25,40,-30,-15), "Sonnet 4.6": (20,-10,-5,20,-10,-5), "Opus 4.6": (0,0,0,0,0,0), "Opus 4.6 Fast": (15,0,0,-35,0,0)}),

    ("Context Window",
     ["32K", "128K", "200K", "500K", "1M"],
     "200K",
     "How much context loaded. 1M beta on Opus 4.6 at standard pricing. Larger=more reference but slower.",
     {"32K": (10,-15,-10,15,-20,0), "128K": (5,-5,-3,5,-5,0), "200K": (0,0,0,0,0,0), "500K": (-8,5,3,-5,8,0), "1M": (-15,10,5,-10,15,0)}),

    ("FILE READING", None, None, None, None),
    ("File Read Depth",
     ["head -30", "head -80", "head -200", "Full file (cat)", "Full + verify"],
     "head -80",
     "How many lines read per file. head -80 = first 80 lines only. Full = entire file. Verify = read + cross-check.",
     {"head -30": (20,-25,-30,10,-30,0), "head -80": (15,-15,-20,5,-15,0), "head -200": (5,-5,-5,0,0,0), "Full file (cat)": (-15,15,20,0,20,0), "Full + verify": (-25,25,30,0,30,0)}),

    ("Binary Files (XLSX/PDF)",
     ["Skip entirely", "Read if asked", "Always extract"],
     "Skip entirely",
     "Whether to read XLSX, PDF, images. Skip = ignore binary files. Always = extract and read content.",
     {"Skip entirely": (10,-15,-20,0,-15,0), "Read if asked": (0,0,0,0,0,0), "Always extract": (-10,15,20,0,15,0)}),

    ("Similar File Reading",
     ["Read 1, skip rest", "Read 3, sample", "Read all"],
     "Read 1, skip rest",
     "When 11 YAMLs or 27 Python files exist, read 1 and assume rest match? Or read all?",
     {"Read 1, skip rest": (15,-20,-25,0,-20,0), "Read 3, sample": (5,-5,-5,0,0,0), "Read all": (-15,20,25,0,20,0)}),

    ("Git Clone Depth",
     ["Shallow (depth 1)", "Full clone", "Full + branches", "Full + log + branches"],
     "Shallow (depth 1)",
     "Shallow = latest commit only. Full = all history. Branches = check all branches. Log = read commit history.",
     {"Shallow (depth 1)": (10,-10,-10,0,-15,0), "Full clone": (0,5,0,0,0,0), "Full + branches": (-5,10,5,0,10,0), "Full + log + branches": (-10,15,10,0,15,0)}),

    ("SEARCH & RESEARCH", None, None, None, None),
    ("Search Queries per Topic",
     ["1-2 (quick)", "3-5 (standard)", "5-10 (thorough)", "10+ (exhaustive)"],
     "1-2 (quick)",
     "How many web searches per topic before responding. More = better coverage but slower.",
     {"1-2 (quick)": (15,-15,-20,0,-15,0), "3-5 (standard)": (5,0,0,0,0,0), "5-10 (thorough)": (-10,15,15,-5,15,0), "10+ (exhaustive)": (-20,20,25,-15,25,0)}),

    ("Web Fetch Depth",
     ["Snippets only", "Fetch when critical", "Always fetch full pages"],
     "Snippets only",
     "Read search result snippets (2-3 lines) or fetch entire web pages for full context.",
     {"Snippets only": (10,-10,-15,0,-10,0), "Fetch when critical": (0,5,5,0,5,0), "Always fetch full pages": (-15,15,20,-10,15,0)}),

    ("Past Chat Coverage",
     ["Skip (current only)", "2-3 pulls (~40 chats)", "Full (all chats)", "Full + read content"],
     "2-3 pulls (~40 chats)",
     "How many past conversations reviewed. Full = paginate through all 159+. Content = read actual words, not summaries.",
     {"Skip (current only)": (15,-20,-25,0,-20,0), "2-3 pulls (~40 chats)": (5,0,0,0,0,0), "Full (all chats)": (-10,10,15,0,15,0), "Full + read content": (-20,20,25,0,25,0)}),

    ("RESPONSE GENERATION", None, None, None, None),
    ("Understanding Before Responding",
     ["40% (rapid)", "60% (current)", "80% (thorough)", "95% (full analysis)"],
     "60% (current)",
     "How much I understand the problem before I start generating output. 40% = start fast. 95% = read everything first.",
     {"40% (rapid)": (20,-25,-30,0,-20,5), "60% (current)": (10,-10,-10,0,-5,0), "80% (thorough)": (-5,10,10,0,10,0), "95% (full analysis)": (-15,20,25,0,25,0)}),

    ("Confidence Language",
     ["Always confident", "Flag unknowns", "Signal uncertainty", "Qualify everything"],
     "Always confident",
     "How I express certainty. Confident = assertive even on weak claims. Signal = explicit about what I know vs don't.",
     {"Always confident": (0,0,-15,0,0,0), "Flag unknowns": (0,5,5,0,0,0), "Signal uncertainty": (0,10,15,0,0,-5), "Qualify everything": (0,5,20,0,0,-10)}),

    ("Build vs Read Priority",
     ["Build first", "Balanced", "Read first", "Read everything first"],
     "Build first",
     "Do I start building immediately or read all context first? Build first = fast but may build wrong thing.",
     {"Build first": (15,-20,-25,0,-15,0), "Balanced": (5,0,0,0,0,0), "Read first": (-10,15,15,0,15,0), "Read everything first": (-20,25,25,0,25,0)}),

    ("CODE GENERATION", None, None, None, None),
    ("Code Verification",
     ["Assume works", "Run once", "Run + fix", "Run + fix + lint + test"],
     "Assume works",
     "Whether I run generated code to verify it works before delivering.",
     {"Assume works": (15,-20,-25,0,0,0), "Run once": (5,0,5,0,0,0), "Run + fix": (-5,15,15,0,0,0), "Run + fix + lint + test": (-15,25,25,-10,0,0)}),

    ("Error Handling in Code",
     ["Skip (add later)", "Basic try/catch", "Comprehensive", "Production + logging"],
     "Skip (add later)",
     "How robust the generated code is on first pass.",
     {"Skip (add later)": (10,-15,0,0,0,0), "Basic try/catch": (5,-5,0,0,0,0), "Comprehensive": (-5,10,5,0,0,0), "Production + logging": (-10,20,10,0,5,0)}),

    ("QUALITY ASSURANCE", None, None, None, None),
    ("QA Before Delivery",
     ["None", "Self-review once", "Verify claims", "Full audit"],
     "None",
     "Whether I review my own output before delivering. Audit = re-read, fact-check, verify all claims.",
     {"None": (15,-20,-20,0,0,0), "Self-review once": (5,5,5,0,0,0), "Verify claims": (-10,15,15,-5,0,0), "Full audit": (-20,25,30,-10,0,0)}),

    ("Hallucination Prevention",
     ["Trust memory", "Flag unverified", "Verify all facts", "Verify + cite"],
     "Trust memory",
     "How I handle factual claims. Trust = use memory as-is. Verify = check every claim against sources.",
     {"Trust memory": (10,-10,-20,0,0,0), "Flag unverified": (5,0,5,0,0,0), "Verify all facts": (-10,10,15,-5,0,0), "Verify + cite": (-15,15,25,-10,0,0)}),

    ("TOOL & OUTPUT", None, None, None, None),
    ("Command Output Reading",
     ["tail -3", "tail -10", "Full output", "Full + parse errors"],
     "tail -3",
     "How much bash/tool output I read. tail -3 = last 3 lines only. Full = everything including warnings.",
     {"tail -3": (10,-10,-15,0,0,0), "tail -10": (5,-5,-5,0,0,0), "Full output": (-5,5,10,0,5,0), "Full + parse errors": (-8,15,15,0,10,0)}),

    ("Tool Calls per Task",
     ["Minimum (1-3)", "Standard (3-5)", "Thorough (5-10)", "Exhaustive (10+)"],
     "Minimum (1-3)",
     "How many tool calls I make per task. More = more data gathered but slower response.",
     {"Minimum (1-3)": (15,-10,-15,0,-10,0), "Standard (3-5)": (5,0,0,0,0,0), "Thorough (5-10)": (-10,10,10,-5,10,0), "Exhaustive (10+)": (-20,15,20,-15,20,0)}),
]

# ─── WRITE SETTINGS ───
row = 3
setting_rows = {}  # track which rows have dropdowns for formulas

for item in settings:
    if item[1] is None:  # Category header
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = cat_fill
            cell.border = thin
        ws.cell(row=row, column=1, value=item[0]).font = cat_font
        ws.cell(row=row, column=1).fill = cat_fill
        ws.row_dimensions[row].height = 22
        row += 1
        continue

    name, options, default, desc, impacts = item
    setting_rows[name] = row

    # A: Setting name
    ws.cell(row=row, column=1, value=name).font = setting_font
    ws.cell(row=row, column=1).alignment = wrap
    ws.cell(row=row, column=1).border = thin

    # B: Dropdown
    cell_b = ws.cell(row=row, column=2, value=default)
    cell_b.font = dropdown_font
    cell_b.fill = dropdown_fill
    cell_b.alignment = center
    cell_b.border = thin
    
    # Add data validation (dropdown)
    options_str = ",".join(options)
    dv = DataValidation(type="list", formula1=f'"{options_str}"', allow_blank=False)
    dv.error = "Pick from the list"
    dv.errorTitle = "Invalid"
    dv.prompt = f"Default: {default}"
    dv.promptTitle = name
    ws.add_data_validation(dv)
    dv.add(cell_b)

    # C: Description
    ws.cell(row=row, column=3, value=desc).font = body_font
    ws.cell(row=row, column=3).alignment = wrap
    ws.cell(row=row, column=3).border = thin

    # D-I: Impact scores per metric (using SWITCH formulas)
    metrics_order = ["speed", "quality", "accuracy", "cost", "depth", "creativity"]
    for mi, metric_idx in enumerate(range(4, 10)):
        # Build a nested IF formula
        parts = []
        for opt in options:
            val = impacts[opt][mi]
            escaped = opt.replace('"', '""')
            parts.append(f'IF(B{row}="{escaped}",{val}')
        
        # Build nested formula: IF(B=opt1,val1,IF(B=opt2,val2,...,0))
        formula = ""
        for p in parts:
            formula += p + ","
        formula += "0" + ")" * len(parts)
        formula = "=" + formula
        
        cell = ws.cell(row=row, column=metric_idx, value=formula)
        cell.font = Font(name="JetBrains Mono", size=9, color="333333")
        cell.alignment = center
        cell.border = thin
        cell.number_format = '0'

    ws.row_dimensions[row].height = 45
    row += 1

# ─── OUTPUT SCORES SECTION ───
row += 1
ws.merge_cells(f"A{row}:I{row}")
ws.cell(row=row, column=1, value="OUTPUT IMPACT SCORES").font = Font(name="JetBrains Mono", size=12, bold=True, color="FFFFFF")
ws.cell(row=row, column=1).fill = PatternFill(start_color="050810", end_color="050810", fill_type="solid")
ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 30
row += 1

# Baselines
baselines = {"SPEED": 65, "QUALITY": 55, "ACCURACY": 45, "COST EFFICIENCY": 50, "DEPTH": 40, "CREATIVITY": 60}
fills = [meter_fill_green, meter_fill_blue, meter_fill_red, meter_fill_amber, meter_fill_cyan, meter_fill_purple]
col_indices = [4, 5, 6, 7, 8, 9]  # D through I

# Find all setting rows (non-category rows with impacts)
setting_data_rows = [r for name, r in setting_rows.items()]

# Headers for output section
ws.cell(row=row, column=1, value="METRIC").font = hdr_font
ws.cell(row=row, column=1).fill = hdr_fill
ws.cell(row=row, column=1).border = thin
ws.cell(row=row, column=2, value="BASELINE").font = hdr_font
ws.cell(row=row, column=2).fill = hdr_fill
ws.cell(row=row, column=2).border = thin
ws.cell(row=row, column=2).alignment = center
ws.cell(row=row, column=3, value="ADJUSTMENT").font = hdr_font
ws.cell(row=row, column=3).fill = hdr_fill
ws.cell(row=row, column=3).border = thin
ws.cell(row=row, column=3).alignment = center
ws.cell(row=row, column=4, value="FINAL SCORE").font = hdr_font
ws.cell(row=row, column=4).fill = hdr_fill
ws.cell(row=row, column=4).border = thin
ws.cell(row=row, column=4).alignment = center
ws.cell(row=row, column=5, value="RATING").font = hdr_font
ws.cell(row=row, column=5).fill = hdr_fill
ws.cell(row=row, column=5).border = thin
ws.cell(row=row, column=5).alignment = center
ws.row_dimensions[row].height = 22
row += 1

for idx, (metric_name, baseline) in enumerate(baselines.items()):
    col_letter = get_column_letter(col_indices[idx])
    
    # A: Metric name
    ws.cell(row=row, column=1, value=metric_name).font = score_label_font
    ws.cell(row=row, column=1).fill = fills[idx]
    ws.cell(row=row, column=1).font = Font(name="JetBrains Mono", size=11, bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).border = thin
    
    # B: Baseline
    ws.cell(row=row, column=2, value=baseline).font = Font(name="JetBrains Mono", size=12, bold=True, color="1E40AF")
    ws.cell(row=row, column=2).alignment = center
    ws.cell(row=row, column=2).border = thin
    
    # C: Adjustment (SUM of all impacts for this metric column)
    sum_parts = [f"{col_letter}{r}" for r in setting_data_rows]
    sum_formula = f"=SUM({','.join(sum_parts)})"
    ws.cell(row=row, column=3, value=sum_formula).font = Font(name="JetBrains Mono", size=12, bold=True)
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=3).border = thin
    ws.cell(row=row, column=3).number_format = '+0;-0;0'
    
    # D: Final Score = MIN(100, MAX(0, baseline + adjustment))
    final_formula = f"=MIN(100,MAX(0,B{row}+C{row}))"
    ws.cell(row=row, column=4, value=final_formula).font = Font(name="JetBrains Mono", size=16, bold=True, color="000000")
    ws.cell(row=row, column=4).alignment = center
    ws.cell(row=row, column=4).border = thin
    ws.cell(row=row, column=4).number_format = '0"%"'
    
    # E: Rating
    rating_formula = f'=IF(D{row}>=80,"EXCELLENT",IF(D{row}>=65,"GOOD",IF(D{row}>=50,"FAIR",IF(D{row}>=35,"WEAK","POOR"))))'
    ws.cell(row=row, column=5, value=rating_formula).font = Font(name="JetBrains Mono", size=10, bold=True)
    ws.cell(row=row, column=5).alignment = center
    ws.cell(row=row, column=5).border = thin
    
    # Conditional formatting on rating
    ws.conditional_formatting.add(f"E{row}", CellIsRule(operator="equal", formula=['"EXCELLENT"'], fill=PatternFill(start_color="065F46", end_color="065F46", fill_type="solid"), font=Font(color="FFFFFF", bold=True)))
    ws.conditional_formatting.add(f"E{row}", CellIsRule(operator="equal", formula=['"GOOD"'], fill=PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid"), font=Font(color="FFFFFF", bold=True)))
    ws.conditional_formatting.add(f"E{row}", CellIsRule(operator="equal", formula=['"FAIR"'], fill=PatternFill(start_color="92400E", end_color="92400E", fill_type="solid"), font=Font(color="FFFFFF", bold=True)))
    ws.conditional_formatting.add(f"E{row}", CellIsRule(operator="equal", formula=['"WEAK"'], fill=PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid"), font=Font(color="FFFFFF", bold=True)))
    ws.conditional_formatting.add(f"E{row}", CellIsRule(operator="equal", formula=['"POOR"'], fill=PatternFill(start_color="450A0A", end_color="450A0A", fill_type="solid"), font=Font(color="FFFFFF", bold=True)))
    
    ws.row_dimensions[row].height = 30
    row += 1

# ─── OVERALL SCORE ───
row += 1
ws.merge_cells(f"A{row}:C{row}")
ws.cell(row=row, column=1, value="OVERALL COMPOSITE SCORE").font = Font(name="JetBrains Mono", size=12, bold=True, color="3B82F6")
ws.cell(row=row, column=1).border = thin
score_rows = list(range(row - 7, row - 1))
avg_formula = f"=ROUND(AVERAGE(D{score_rows[0]},D{score_rows[1]},D{score_rows[2]},D{score_rows[3]},D{score_rows[4]},D{score_rows[5]}),0)"
ws.cell(row=row, column=4, value=avg_formula).font = Font(name="JetBrains Mono", size=20, bold=True, color="3B82F6")
ws.cell(row=row, column=4).alignment = center
ws.cell(row=row, column=4).border = thin
ws.cell(row=row, column=4).number_format = '0"%"'
overall_rating = f'=IF(D{row}>=80,"EXCELLENT",IF(D{row}>=65,"GOOD",IF(D{row}>=50,"FAIR",IF(D{row}>=35,"WEAK","POOR"))))'
ws.cell(row=row, column=5, value=overall_rating).font = Font(name="JetBrains Mono", size=12, bold=True)
ws.cell(row=row, column=5).alignment = center
ws.cell(row=row, column=5).border = thin
ws.conditional_formatting.add(f"E{row}", CellIsRule(operator="equal", formula=['"EXCELLENT"'], fill=PatternFill(start_color="065F46", end_color="065F46", fill_type="solid"), font=Font(color="FFFFFF", bold=True)))
ws.conditional_formatting.add(f"E{row}", CellIsRule(operator="equal", formula=['"GOOD"'], fill=PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid"), font=Font(color="FFFFFF", bold=True)))
ws.conditional_formatting.add(f"E{row}", CellIsRule(operator="equal", formula=['"FAIR"'], fill=PatternFill(start_color="92400E", end_color="92400E", fill_type="solid"), font=Font(color="FFFFFF", bold=True)))

ws.row_dimensions[row].height = 35

# ─── INSTRUCTIONS ───
row += 2
ws.merge_cells(f"A{row}:I{row}")
ws.cell(row=row, column=1, value="HOW TO USE: Click any yellow cell in column B. A dropdown appears. Change the setting. Columns D-I update automatically showing impact on each metric. Output scores at the bottom recalculate in real time. RATING column shows EXCELLENT/GOOD/FAIR/WEAK/POOR.").font = Font(name="Arial", size=9, color="666666", italic=True)
ws.cell(row=row, column=1).alignment = wrap
ws.row_dimensions[row].height = 35

# ═══════════════════════════════════════════
# Recalc with LibreOffice
# ═══════════════════════════════════════════
output_path = "/mnt/user-data/outputs/CLAUDE_SETTINGS_CONFIGURATOR.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Settings with dropdowns: {len(setting_rows)}")
print(f"Output metrics: {len(baselines)}")
